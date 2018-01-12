##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import functools
from unittest import mock

import faker
import openpyxl
from django.contrib.auth.models import Permission, User
from django.test import TestCase

from internship.models.organization import Organization
from internship.tests.factories.cohort import CohortFactory
from internship.views.upload_xls import _save_xls_place as save_xls_place
from reference.tests.factories import country

faker = faker.Faker()


class XlsPlaceImportTestCase(TestCase):
    def setUp(self):
        self.cohort = CohortFactory()
        self.user = User.objects.create_user('demo',
                                             email='demo@demo.org',
                                             password='password')
        permission = Permission.objects.get(codename='is_internship_manager')
        self.user.user_permissions.add(permission)
        self.user.save()

    @classmethod
    def generate_workbook(cls):
        col_reference, col_name, col_address, \
            col_postal_code, col_city, col_country, \
            col_url = range(0, 7)

        columns = (
            (col_reference, lambda: faker.random_int(1, 500)),
            (col_name, lambda: 'Hospital {}'.format(faker.name())),
            (col_address, lambda: faker.address().replace('\n', '')),
            (col_postal_code, faker.postalcode),
            (col_city, faker.city),
            (col_country, lambda: country.CountryFactory().iso_code),
            (col_url, faker.url)
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for row in range(1, 11):
            # add_cell = functools.partial(worksheet.cell, row=row)
            for column, func in columns:
                worksheet.cell(row=row, column=column+1).value = func()
                # add_cell(column=column+1, value=func())

        return workbook

    def test_import_xls_from_python_api(self):
        with mock.patch('openpyxl.load_workbook') as mock_load_workbook:
            workbook = self.generate_workbook()
            mock_load_workbook.return_value = workbook

            qs = Organization.objects.all()
            self.assertEqual(qs.count(), 0)
            save_xls_place(None, cohort=self.cohort)

            self.assertTrue(mock_load_workbook.called)
            qs = Organization.objects.all()
            self.assertTrue(qs.count() > 0)
